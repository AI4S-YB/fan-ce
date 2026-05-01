"""
Migrate rose_phenotype_test.xlsx (wide format) → trial/plot/observation SQLite.

Wide columns like "2021年花瓣数量" are decomposed into:
  trait_code="花瓣数量", timepoint="2021"
"""
import sqlite3
import os
import re

import openpyxl

SRC = os.environ.get('PHENOME_SRC', 'test_data/phenotype/rose_phenotype_test.xlsx')
DST = os.environ.get('PHENOME_DST', 'test_data/phenotype/rose_phenotype_v2.db')

# ---- Read Excel ----
wb = openpyxl.load_workbook(SRC)
ws = wb.active
header = [str(c.value or '') for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"Header ({len(header)}): {header}")

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append([v for v in row])

print(f"Data rows: {len(rows)}")

# ---- Classify columns ----
ID_COL_IDX = 0
NAME_CN_IDX = 1
NAME_EN_IDX = 2

# Columns 3..16 are traits
trait_cols = list(range(3, len(header)))

YEAR_PATTERN = re.compile(r'^(\d{4})年(.+)$')

def parse_trait_col(col_name: str):
    """Return (trait_code, timepoint) from a column name."""
    m = YEAR_PATTERN.match(col_name)
    if m:
        return m.group(2).strip(), m.group(1)  # trait_code, year as timepoint
    return col_name.strip(), ''

# ---- Build SQLite ----
if os.path.exists(DST):
    os.remove(DST)

db = sqlite3.connect(DST)
db.execute("PRAGMA journal_mode=WAL;")

db.execute("""
CREATE TABLE trial (
    id          INTEGER PRIMARY KEY,
    trial_name  TEXT,
    location    TEXT,
    year        INTEGER,
    season      TEXT,
    trial_type  TEXT,
    design_type TEXT,
    remark      TEXT
);
""")

db.execute("""
CREATE TABLE plot (
    id            INTEGER PRIMARY KEY,
    trial_id      INTEGER REFERENCES trial(id),
    germplasm_id  INTEGER,
    plot_code     TEXT,
    subject_name  TEXT,
    subject_name_cn TEXT,
    subject_name_en TEXT,
    block         INTEGER,
    rep           INTEGER,
    row           INTEGER,
    col           INTEGER,
    treatment     TEXT
);
""")

db.execute("""
CREATE TABLE observation (
    id              INTEGER PRIMARY KEY,
    plot_id         INTEGER REFERENCES plot(id),
    trait_code      TEXT,
    value_numeric   REAL,
    value_text      TEXT,
    value_category  TEXT,
    timepoint       TEXT,
    obs_date        TEXT,
    is_missing      INTEGER DEFAULT 0
);
""")

# ---- Insert trial ----
db.execute("""
INSERT INTO trial (id, trial_name, location, year, season, trial_type, design_type, remark)
VALUES (1, ?, '昆明', 2021, '春季', '品种比较', '完全随机',
        '示例数据，基于公开月季品种信息生成');
""", ("月季品种评价试验",))

# ---- Insert plots & observations ----
for row_idx, row_data in enumerate(rows):
    subject_id = str(row_data[ID_COL_IDX] or '').strip()
    if not subject_id:
        continue

    name_cn = str(row_data[NAME_CN_IDX] or '').strip()
    name_en = str(row_data[NAME_EN_IDX] or '').strip()

    db.execute("""
    INSERT INTO plot (id, trial_id, plot_code, subject_name, subject_name_cn, subject_name_en, block, rep)
    VALUES (?, 1, ?, ?, ?, ?, 1, ?);
    """, (row_idx + 1, subject_id, name_en or name_cn, name_cn, name_en, row_idx + 1))

    plot_id = row_idx + 1

    for col_idx in trait_cols:
        col_name = header[col_idx]
        raw_value = row_data[col_idx]
        trait_code, timepoint = parse_trait_col(col_name)

        # Determine value type
        value_str = str(raw_value or '').strip()
        is_missing = 0
        value_numeric = None
        value_text = None

        if value_str in ('', 'NA', 'N/A', 'null', 'None'):
            is_missing = 1
        else:
            try:
                value_numeric = float(value_str)
            except ValueError:
                value_text = value_str

        db.execute("""
        INSERT INTO observation (plot_id, trait_code, value_numeric, value_text, timepoint, is_missing)
        VALUES (?, ?, ?, ?, ?, ?);
        """, (plot_id, trait_code, value_numeric, value_text, timepoint, is_missing))

db.commit()

# ---- Verify ----
print(f"\ntrial:      {db.execute('SELECT COUNT(*) FROM trial').fetchone()[0]}")
print(f"plot:       {db.execute('SELECT COUNT(*) FROM plot').fetchone()[0]}")
print(f"observation:{db.execute('SELECT COUNT(*) FROM observation').fetchone()[0]}")

print("\nSample observations:")
for row in db.execute("""
    SELECT p.plot_code, p.subject_name_cn, o.trait_code, o.value_numeric, o.value_text, o.timepoint
    FROM observation o JOIN plot p ON p.id = o.plot_id
    ORDER BY p.id, o.trait_code LIMIT 15
"""):
    print(row)

db.close()
print(f"\nWrote: {DST}")
