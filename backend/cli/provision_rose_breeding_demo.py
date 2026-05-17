from __future__ import annotations

import gzip
import json
import os
import sqlite3
from collections import OrderedDict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import or_

from modules.breeding.models import (
    BreedingAssay,
    BreedingBioSample,
    BreedingDataFile,
    BreedingDatasetAssayLink,
    BreedingDatasetSubjectLink,
    BreedingMaterial,
    BreedingObservation,
    BreedingPhenotypeSubjectMap,
    BreedingPlot,
    BreedingProgram,
    BreedingTrial,
    BreedingVariantSampleMap,
)
from modules.datasets.models import AssetFile, DatasetAsset, DatasetVersion
from shared.database import MyDBManager


ROOT_DIR = Path(os.environ.get("TEST_DATA_DIR", Path(__file__).resolve().parent.parent / "test_data"))
GERMPLASM_XLSX = ROOT_DIR / "germplasm" / "rose_germplasm_test.xlsx"
PHENOTYPE_DB = ROOT_DIR / "phenotype" / "rose_phenotype_test.db"

PROGRAM_CODE = "rose-breeding-demo"
PROGRAM_NAME = "月季观赏性与采后联合评价项目"
REPRESENTATIVE_ACCESSIONS = ["RH00004", "RH00012", "RH00027", "RH00039"]

TRAIT_SPECS = {
    "花瓣长": ("petal_length", date(2023, 5, 20)),
    "花瓣宽": ("petal_width", date(2023, 5, 20)),
    "花瓣长宽比值": ("petal_length_width_ratio", date(2023, 5, 20)),
    "多头数": ("cluster_flower_count", date(2023, 5, 20)),
    "连续开花性": ("continuous_flowering_score", date(2023, 5, 20)),
    "乙烯敏感脱落": ("ethylene_abscission_score", date(2023, 5, 20)),
    "瓶插寿命": ("vase_life_days", date(2023, 5, 20)),
    "乙烯敏感衰老": ("ethylene_senescence_score", date(2023, 5, 20)),
    "2021年花瓣数量": ("petal_count_2021", date(2021, 5, 20)),
    "2022年花瓣数量": ("petal_count_2022", date(2022, 5, 20)),
    "2023年花瓣数量": ("petal_count_2023", date(2023, 5, 20)),
    "2021年花直径": ("flower_diameter_2021", date(2021, 5, 20)),
    "2022年花直径": ("flower_diameter_2022", date(2022, 5, 20)),
    "2023年花直径": ("flower_diameter_2023", date(2023, 5, 20)),
}


def load_germplasm_rows() -> OrderedDict[str, dict]:
    workbook = load_workbook(GERMPLASM_XLSX, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(item) if item is not None else "" for item in next(rows)]
    result: OrderedDict[str, dict] = OrderedDict()
    for row in rows:
        item = {header[index]: value for index, value in enumerate(row)}
        accession = str(item.get("ID") or "").strip()
        if accession:
            result[accession] = item
    return result


def load_phenotype_rows() -> list[dict]:
    conn = sqlite3.connect(PHENOTYPE_DB)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute('SELECT * FROM phenotype ORDER BY ID').fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def read_variant_samples(variant_file_path: str) -> list[str]:
    with gzip.open(variant_file_path, "rt", encoding="utf-8") as handle:
        for line in handle:
            if line.startswith("#CHROM"):
                return line.rstrip("\n").split("\t")[9:]
    return []


def delete_existing_program(db, program: BreedingProgram):
    material_ids = [item[0] for item in db.query(BreedingMaterial.id).filter(BreedingMaterial.program_id == program.id).all()]
    trial_ids = [item[0] for item in db.query(BreedingTrial.id).filter(BreedingTrial.program_id == program.id).all()]
    plot_ids = [item[0] for item in db.query(BreedingPlot.id).filter(BreedingPlot.trial_id.in_(trial_ids)).all()] if trial_ids else []
    biosample_ids = [item[0] for item in db.query(BreedingBioSample.id).filter(BreedingBioSample.material_id.in_(material_ids)).all()] if material_ids else []
    assay_ids = [item[0] for item in db.query(BreedingAssay.id).filter(BreedingAssay.biosample_id.in_(biosample_ids)).all()] if biosample_ids else []

    if assay_ids:
        db.query(BreedingDataFile).filter(BreedingDataFile.assay_id.in_(assay_ids)).delete(synchronize_session=False)
        db.query(BreedingDatasetAssayLink).filter(BreedingDatasetAssayLink.assay_id.in_(assay_ids)).delete(synchronize_session=False)
    if material_ids or biosample_ids or plot_ids:
        db.query(BreedingVariantSampleMap).filter(
            or_(
                BreedingVariantSampleMap.material_id.in_(material_ids or [-1]),
                BreedingVariantSampleMap.biosample_id.in_(biosample_ids or [-1]),
                BreedingVariantSampleMap.plot_id.in_(plot_ids or [-1]),
            )
        ).delete(synchronize_session=False)
        db.query(BreedingPhenotypeSubjectMap).filter(
            or_(
                BreedingPhenotypeSubjectMap.material_id.in_(material_ids or [-1]),
                BreedingPhenotypeSubjectMap.plot_id.in_(plot_ids or [-1]),
                BreedingPhenotypeSubjectMap.trial_id.in_(trial_ids or [-1]),
            )
        ).delete(synchronize_session=False)
        db.query(BreedingDatasetSubjectLink).filter(
            or_(
                BreedingDatasetSubjectLink.program_id == program.id,
                BreedingDatasetSubjectLink.material_id.in_(material_ids or [-1]),
                BreedingDatasetSubjectLink.trial_id.in_(trial_ids or [-1]),
                BreedingDatasetSubjectLink.plot_id.in_(plot_ids or [-1]),
                BreedingDatasetSubjectLink.biosample_id.in_(biosample_ids or [-1]),
            )
        ).delete(synchronize_session=False)
        db.query(BreedingObservation).filter(BreedingObservation.trial_id.in_(trial_ids or [-1])).delete(synchronize_session=False)
    if assay_ids:
        db.query(BreedingAssay).filter(BreedingAssay.id.in_(assay_ids)).delete(synchronize_session=False)
    if biosample_ids:
        db.query(BreedingBioSample).filter(BreedingBioSample.id.in_(biosample_ids)).delete(synchronize_session=False)
    if plot_ids:
        db.query(BreedingPlot).filter(BreedingPlot.id.in_(plot_ids)).delete(synchronize_session=False)
    if trial_ids:
        db.query(BreedingTrial).filter(BreedingTrial.id.in_(trial_ids)).delete(synchronize_session=False)
    if material_ids:
        db.query(BreedingMaterial).filter(BreedingMaterial.id.in_(material_ids)).delete(synchronize_session=False)
    db.query(BreedingProgram).filter(BreedingProgram.id == program.id).delete(synchronize_session=False)
    db.flush()


def resolve_current_version(db, fragment: str, dataset_type: str) -> DatasetVersion:
    version = (
        db.query(DatasetVersion)
        .filter(
            DatasetVersion.is_current == 1,
            DatasetVersion.dataset_type == dataset_type,
            DatasetVersion.file_path.like(f"%{fragment}%"),
        )
        .order_by(DatasetVersion.id.asc())
        .first()
    )
    if version is None:
        raise RuntimeError(f"current dataset version not found for {dataset_type} with fragment {fragment}")
    return version


def resolve_assets(db, version_id: int) -> dict[str, DatasetAsset]:
    assets = db.query(DatasetAsset).filter(DatasetAsset.dataset_version_id == version_id).all()
    return {str(asset.asset_code): asset for asset in assets}


def resolve_asset_files(db, asset_id: int) -> list[AssetFile]:
    return (
        db.query(AssetFile)
        .filter(AssetFile.dataset_asset_id == asset_id)
        .order_by(AssetFile.id.asc())
        .all()
    )


def create_program(db) -> BreedingProgram:
    program = BreedingProgram(
        code=PROGRAM_CODE,
        name=PROGRAM_NAME,
        species_name="Rosa chinensis",
        breeding_goal=(
            "围绕月季观赏品质、连续开花能力、瓶插寿命与乙烯响应开展联合评价，"
            "形成可关联表型、变异组与转录组数据的展示型 breeding 示例项目。"
        ),
        start_year=2021,
        status="active",
        owner_name="FAN-CE inferred curator",
        meta_json=json.dumps(
            {
                "source_files": {
                    "germplasm": str(GERMPLASM_XLSX),
                    "phenotype": str(PHENOTYPE_DB),
                },
                "assumptions": [
                    "基于表型宽表和种质谱系信息推断单项目单试验结构",
                    "RNA 与 DNA 样本仅为展示型代表样本，不代表真实实验设计全量还原",
                ],
            },
            ensure_ascii=False,
        ),
        created_by=1,
        updated_by=1,
    )
    db.add(program)
    db.flush()
    return program


def main():
    germplasm_rows = load_germplasm_rows()
    phenotype_rows = load_phenotype_rows()

    material_rows = [row for row in phenotype_rows if str(row.get("ID") or "") in germplasm_rows]
    if not material_rows:
        raise RuntimeError("no overlapping accessions between phenotype and germplasm")

    with MyDBManager() as db:
        existing_program = db.query(BreedingProgram).filter(BreedingProgram.code == PROGRAM_CODE).first()
        if existing_program:
            delete_existing_program(db, existing_program)

        transcriptome_version = resolve_current_version(db, "/test_data/transcriptome/", "transcriptome")
        variome_version = resolve_current_version(db, "/test_data/variome/", "variome")
        phenome_version = resolve_current_version(db, "/test_data/phenotype/", "phenome")

        transcriptome_assets = resolve_assets(db, transcriptome_version.id)
        variome_assets = resolve_assets(db, variome_version.id)
        phenome_assets = resolve_assets(db, phenome_version.id)

        expression_asset = transcriptome_assets["expression_matrix"]
        sample_meta_asset = transcriptome_assets["sample_metadata"]
        count_raw_asset = transcriptome_assets["count_matrix_raw"]
        fpkm_raw_asset = transcriptome_assets["fpkm_matrix_raw"]
        variant_asset = next(iter(variome_assets.values()))
        phenotype_asset = phenome_assets["phenotype_index"]

        expression_files = resolve_asset_files(db, expression_asset.id)
        sample_meta_files = resolve_asset_files(db, sample_meta_asset.id)
        count_raw_files = resolve_asset_files(db, count_raw_asset.id)
        fpkm_raw_files = resolve_asset_files(db, fpkm_raw_asset.id)
        variant_files = resolve_asset_files(db, variant_asset.id)

        program = create_program(db)

        trial = BreedingTrial(
            program_id=program.id,
            trial_code="ROSE-TRIAL-2021-2023",
            trial_name="月季花部性状与采后表现联合评价",
            trial_type="field",
            objective="对核心月季种质开展 2021-2023 连续观测，构建花部性状、瓶插寿命和乙烯响应的结构化展示样本。",
            location_name="月季资源圃示范评价点",
            season_label="2021-2023",
            design_type="single-plot comparative",
            replicate_count=1,
            status="active",
            sowing_date=date(2021, 3, 1),
            harvest_date=date(2023, 10, 30),
            meta_json=json.dumps({"source_dataset_version_id": phenome_version.id}, ensure_ascii=False),
            created_by=1,
            updated_by=1,
        )
        db.add(trial)
        db.flush()

        material_by_accession: dict[str, BreedingMaterial] = {}
        plot_by_accession: dict[str, BreedingPlot] = {}

        for index, row in enumerate(material_rows, start=1):
            accession = str(row["ID"])
            germplasm = germplasm_rows[accession]
            material = BreedingMaterial(
                program_id=program.id,
                material_code=accession,
                material_name=str(germplasm.get("chinese_name") or accession),
                material_type="cultivar",
                generation_code="pedigree_known",
                origin=f"{germplasm.get('P_name') or '未知父本'} × {germplasm.get('M_name') or '未知母本'}",
                germplasm_accession=accession,
                germplasm_name=str(germplasm.get("chinese_name") or accession),
                germplasm_source_file=str(GERMPLASM_XLSX),
                status="active",
                is_check=1 if accession in {"RH00033", "RH00039"} else 0,
                meta_json=json.dumps(
                    {
                        "english_name": germplasm.get("english_name"),
                        "father_accession": germplasm.get("P"),
                        "mother_accession": germplasm.get("M"),
                        "father_name": germplasm.get("P_name"),
                        "mother_name": germplasm.get("M_name"),
                        "history": germplasm.get("育种历史"),
                        "flower_traits": germplasm.get("花朵性状"),
                        "plant_traits": germplasm.get("植株特征"),
                        "usage": germplasm.get("用途"),
                    },
                    ensure_ascii=False,
                ),
                remarks="由 germplasm 与 phenotype 样例数据联合推断的 breeding 展示材料。",
                created_by=1,
                updated_by=1,
            )
            db.add(material)
            db.flush()
            material_by_accession[accession] = material

            plot = BreedingPlot(
                trial_id=trial.id,
                material_id=material.id,
                plot_code=f"PLOT-{accession}",
                replicate_no=1,
                block_no=1,
                row_no=index,
                col_no=1,
                treatment_code="standard_field",
                area=1.0,
                plant_count_planned=3,
                plant_count_actual=3,
                status="active",
                meta_json=json.dumps({"source_subject_id": accession}, ensure_ascii=False),
                remarks="按表型主体一对一生成的展示型 plot。",
                created_by=1,
                updated_by=1,
            )
            db.add(plot)
            db.flush()
            plot_by_accession[accession] = plot

            db.add(
                BreedingDatasetSubjectLink(
                    dataset_id=phenome_version.database_id,
                    version_id=phenome_version.id,
                    asset_id=phenotype_asset.id,
                    material_id=material.id,
                    role="phenotype_subject",
                    mapping_status="reviewed",
                    mapping_method="inferred",
                    confidence="high",
                    is_primary=1,
                    note="根据 phenotype 行 ID 与 germplasm accession 一致直接挂接。",
                    created_by=1,
                    updated_by=1,
                )
            )
            db.add(
                BreedingPhenotypeSubjectMap(
                    dataset_id=phenome_version.database_id,
                    version_id=phenome_version.id,
                    asset_id=phenotype_asset.id,
                    row_key=accession,
                    trial_id=trial.id,
                    plot_id=plot.id,
                    material_id=material.id,
                    trait_code="panel_row",
                    obs_date=date(2023, 5, 20),
                    mapping_status="reviewed",
                    mapping_method="inferred",
                    confidence="high",
                    is_primary=1,
                    note="表型宽表主体行到 breeding material/plot 的主映射。",
                    created_by=1,
                    updated_by=1,
                )
            )

        db.add(
            BreedingDatasetSubjectLink(
                dataset_id=phenome_version.database_id,
                version_id=phenome_version.id,
                asset_id=phenotype_asset.id,
                trial_id=trial.id,
                role="phenotype_panel",
                mapping_status="reviewed",
                mapping_method="inferred",
                confidence="high",
                is_primary=1,
                note="月季表型宽表整体对应当前 trial 的评价面板。",
                created_by=1,
                updated_by=1,
            )
        )

        observation_count = 0
        for row in material_rows:
            accession = str(row["ID"])
            material = material_by_accession[accession]
            plot = plot_by_accession[accession]
            for column_name, value in row.items():
                if column_name == "ID" or value in (None, ""):
                    continue
                if column_name not in TRAIT_SPECS:
                    continue
                trait_code, obs_date = TRAIT_SPECS[column_name]
                db.add(
                    BreedingObservation(
                        trial_id=trial.id,
                        plot_id=plot.id,
                        material_id=material.id,
                        observation_level="plot",
                        trait_code=trait_code,
                        trait_name=column_name,
                        protocol_name="rose_phenotype_demo_projection",
                        obs_value_num=float(value),
                        obs_date=obs_date,
                        observer="FAN-CE inferred curator",
                        qc_status="reviewed",
                        source_dataset_id=phenome_version.database_id,
                        source_version_id=phenome_version.id,
                        source_asset_id=phenotype_asset.id,
                        source_row_key=accession,
                        meta_json=json.dumps({"source_column": column_name}, ensure_ascii=False),
                        remarks="由 phenotype sqlite 宽表自动投影到 breeding observation。",
                        created_by=1,
                        updated_by=1,
                    )
                )
                observation_count += 1

        representative_materials = [material_by_accession[item] for item in REPRESENTATIVE_ACCESSIONS if item in material_by_accession]
        representative_variant_samples = read_variant_samples(variome_version.file_path)[: len(representative_materials)]

        first_rna_assay = None
        first_dna_assay = None

        for index, material in enumerate(representative_materials, start=1):
            accession = material.material_code
            plot = plot_by_accession[accession]

            rna_biosample = BreedingBioSample(
                sample_code=f"BS-{accession}-RNA",
                material_id=material.id,
                plot_id=plot.id,
                sample_type="RNA",
                tissue_type="petal",
                timepoint="full_bloom",
                treatment_label="standard_field",
                collection_date=date(2023, 5, 18),
                collector="FAN-CE inferred curator",
                storage_location="LN2-Rose-RNA",
                status="active",
                remarks="展示型 RNA 样本，用于承接 transcriptome dataset。",
                created_by=1,
                updated_by=1,
            )
            db.add(rna_biosample)
            db.flush()

            dna_biosample = BreedingBioSample(
                sample_code=f"BS-{accession}-DNA",
                material_id=material.id,
                plot_id=plot.id,
                sample_type="DNA",
                tissue_type="leaf",
                timepoint="vegetative",
                treatment_label="standard_field",
                collection_date=date(2023, 4, 15),
                collector="FAN-CE inferred curator",
                storage_location="DNA-Rose-Freezer-A",
                status="active",
                remarks="展示型 DNA 样本，用于承接 variome dataset。",
                created_by=1,
                updated_by=1,
            )
            db.add(dna_biosample)
            db.flush()

            rna_assay = BreedingAssay(
                assay_code=f"AS-{accession}-RNASEQ",
                biosample_id=rna_biosample.id,
                assay_type="RNAseq",
                platform="Illumina NovaSeq",
                vendor="inferred-demo-pipeline",
                run_date=date(2024, 1, 15),
                status="active",
                remarks="展示型 RNA-seq assay，关联现有 transcriptome 资产。",
                created_by=1,
                updated_by=1,
            )
            db.add(rna_assay)
            db.flush()

            dna_assay = BreedingAssay(
                assay_code=f"AS-{accession}-GENO",
                biosample_id=dna_biosample.id,
                assay_type="Genotyping",
                platform="VCF panel",
                vendor="inferred-demo-pipeline",
                run_date=date(2024, 2, 1),
                status="active",
                remarks="展示型变异检测 assay，关联现有 variome 资产。",
                created_by=1,
                updated_by=1,
            )
            db.add(dna_assay)
            db.flush()

            if first_rna_assay is None:
                first_rna_assay = rna_assay
            if first_dna_assay is None:
                first_dna_assay = dna_assay

            for asset in (expression_asset,):
                db.add(
                    BreedingDatasetAssayLink(
                        dataset_id=transcriptome_version.database_id,
                        version_id=transcriptome_version.id,
                        asset_id=asset.id,
                        assay_id=rna_assay.id,
                        role="expression_matrix",
                        mapping_status="reviewed",
                        mapping_method="inferred",
                        confidence="medium",
                        is_primary=1 if asset.id == expression_asset.id else 0,
                        note="展示型 assay 与 transcriptome 表达矩阵挂接。",
                        created_by=1,
                        updated_by=1,
                    )
                )
            db.add(
                BreedingDatasetAssayLink(
                    dataset_id=variome_version.database_id,
                    version_id=variome_version.id,
                    asset_id=variant_asset.id,
                    assay_id=dna_assay.id,
                    role="variant_calls",
                    mapping_status="reviewed",
                    mapping_method="inferred",
                    confidence="medium",
                    is_primary=1,
                    note="展示型 assay 与 variome callset 挂接。",
                    created_by=1,
                    updated_by=1,
                )
            )

            if index - 1 < len(representative_variant_samples):
                db.add(
                    BreedingVariantSampleMap(
                        dataset_id=variome_version.database_id,
                        version_id=variome_version.id,
                        asset_id=variant_asset.id,
                        vcf_sample_name=representative_variant_samples[index - 1],
                        normalized_sample_name=accession,
                        sample_alias=f"{accession}-demo",
                        material_id=material.id,
                        biosample_id=dna_biosample.id,
                        plot_id=plot.id,
                        mapping_status="draft",
                        mapping_method="inferred",
                        confidence="low",
                        is_primary=1,
                        note="仅用于展示层示例，基于样本顺序做低置信度占位映射。",
                        created_by=1,
                        updated_by=1,
                    )
                )

        if first_rna_assay is not None:
            for asset, asset_files, role in (
                (expression_asset, expression_files, "expression_matrix"),
                (sample_meta_asset, sample_meta_files, "sample_metadata"),
                (count_raw_asset, count_raw_files, "count_matrix_raw"),
                (fpkm_raw_asset, fpkm_raw_files, "fpkm_matrix_raw"),
            ):
                for asset_file in asset_files:
                    db.add(
                        BreedingDataFile(
                            assay_id=first_rna_assay.id,
                            source_mode="dataset_file",
                            dataset_id=transcriptome_version.database_id,
                            version_id=transcriptome_version.id,
                            asset_id=asset.id,
                            asset_file_id=asset_file.id,
                            file_role=role if asset_file.file_role == "primary" else f"{role}_{asset_file.file_role}",
                            file_name=asset_file.file_name,
                            file_format=asset_file.file_format,
                            uri_snapshot=asset_file.local_path,
                            file_size=asset_file.file_size,
                            status="active",
                            remarks="由 transcriptome dataset 资产文件直接映射。",
                            created_by=1,
                            updated_by=1,
                        )
                    )

        if first_dna_assay is not None:
            for asset_file in variant_files:
                db.add(
                    BreedingDataFile(
                        assay_id=first_dna_assay.id,
                        source_mode="dataset_file",
                        dataset_id=variome_version.database_id,
                        version_id=variome_version.id,
                        asset_id=variant_asset.id,
                        asset_file_id=asset_file.id,
                        file_role="variant_calls" if asset_file.file_role == "primary" else "variant_index",
                        file_name=asset_file.file_name,
                        file_format=asset_file.file_format,
                        uri_snapshot=asset_file.local_path,
                        file_size=asset_file.file_size,
                        status="active",
                        remarks="由 variome dataset 资产文件直接映射。",
                        created_by=1,
                        updated_by=1,
                    )
                )

        db.commit()

        summary = {
            "program_code": program.code,
            "program_id": program.id,
            "materials": len(material_by_accession),
            "plots": len(plot_by_accession),
            "observations": observation_count,
            "representative_assays": len(representative_materials) * 2,
            "variant_sample_maps": len(representative_variant_samples),
            "phenome_dataset_id": phenome_version.database_id,
            "transcriptome_dataset_id": transcriptome_version.database_id,
            "variome_dataset_id": variome_version.database_id,
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
