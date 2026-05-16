from __future__ import annotations

import json
import re
import shutil
import tempfile
import uuid
from pathlib import Path

from openpyxl import load_workbook


VALIDATION_ROOT = Path(tempfile.gettempdir()) / "fan_ce_germplasm_validation"
CANONICAL_TEMPLATE_PROFILE = "crop_germplasm_v1"
TEMPLATE_PROFILE_ALIASES = {
    "crop_germplasm_v1": CANONICAL_TEMPLATE_PROFILE,
    "rose_germplasm_v1": CANONICAL_TEMPLATE_PROFILE,
}
SUPPORTED_TEMPLATE_PROFILES = set(TEMPLATE_PROFILE_ALIASES)
GERMPLASM_REQUIRED_HEADERS = ("ID", "chinese_name")
BUILTIN_FIELD_SPECS = {
    "ID": {"field_key": "accession_id", "field_label": "种质编号", "data_type": "string"},
    "chinese_name": {"field_key": "display_name", "field_label": "显示名称", "data_type": "string"},
    "english_name": {"field_key": "english_name", "field_label": "英文名称", "data_type": "string"},
    "P": {"field_key": "father_accession", "field_label": "父本编号", "data_type": "string"},
    "M": {"field_key": "mother_accession", "field_label": "母本编号", "data_type": "string"},
    "P_name": {"field_key": "father_name_snapshot", "field_label": "父本名称", "data_type": "string"},
    "M_name": {"field_key": "mother_name_snapshot", "field_label": "母本名称", "data_type": "string"},
}


def normalize_template_profile(template_profile: str) -> str:
    normalized = TEMPLATE_PROFILE_ALIASES.get((template_profile or "").strip())
    if normalized is None:
        raise ValueError(f"unsupported template profile: {template_profile}")
    return normalized


def ensure_validation_root() -> Path:
    VALIDATION_ROOT.mkdir(parents=True, exist_ok=True)
    return VALIDATION_ROOT


def persist_uploaded_excel(upload, filename_hint: str | None = None) -> tuple[str, str]:
    ensure_validation_root()
    safe_name = Path(filename_hint or upload.filename or "germplasm.xlsx").name
    suffix = Path(safe_name).suffix or ".xlsx"
    token = uuid.uuid4().hex
    bundle_dir = VALIDATION_ROOT / token
    bundle_dir.mkdir(parents=True, exist_ok=True)
    source_path = bundle_dir / f"source{suffix}"
    upload.file.seek(0)
    with open(source_path, "wb") as output_handle:
        shutil.copyfileobj(upload.file, output_handle)
    return token, str(source_path)


def bundle_manifest_path(token: str) -> Path:
    return ensure_validation_root() / token / "manifest.json"


def write_validation_bundle(token: str, payload: dict) -> None:
    manifest_path = bundle_manifest_path(token)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    with open(manifest_path, "w", encoding="utf-8") as output_handle:
        json.dump(payload, output_handle, ensure_ascii=False, indent=2)


def load_validation_bundle(token: str) -> dict:
    manifest_path = bundle_manifest_path(token)
    if not manifest_path.exists():
        raise FileNotFoundError(f"validation token not found: {token}")
    with open(manifest_path, "r", encoding="utf-8") as input_handle:
        return json.load(input_handle)


def update_validation_bundle(token: str, updates: dict) -> dict:
    payload = load_validation_bundle(token)
    payload.update(updates)
    write_validation_bundle(token, payload)
    return payload


def _clean_cell(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "na", "null"}:
        return None
    return text


def _is_blank_row(row_values) -> bool:
    return all(_clean_cell(value) is None for value in row_values)


def _is_comment_row(row_values) -> bool:
    first_value = next((value for value in row_values if _clean_cell(value) is not None), None)
    first_text = _clean_cell(first_value)
    return bool(first_text and first_text.startswith("#"))


def _load_sheet_headers_and_rows(sheet):
    all_rows = list(sheet.iter_rows(values_only=True))
    header_row_no = None
    headers = []
    data_rows = []

    for excel_row_no, row_values in enumerate(all_rows, start=1):
        if _is_blank_row(row_values):
            continue
        if _is_comment_row(row_values):
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in row_values]
        header_row_no = excel_row_no
        break

    if header_row_no is None:
        return 1, [], []

    for excel_row_no, row_values in enumerate(all_rows[header_row_no:], start=header_row_no + 1):
        if _is_blank_row(row_values):
            continue
        if _is_comment_row(row_values):
            continue
        data_rows.append((excel_row_no, row_values))
    return header_row_no, headers, data_rows


def _validate_germplasm_headers(headers: list[str], *, header_row_no: int) -> tuple[list[dict], list[dict]]:
    errors: list[dict] = []
    warnings: list[dict] = []
    seen_headers: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        if not header:
            continue
        if header in seen_headers:
            errors.append(
                {
                    "row_no": header_row_no,
                    "column_name": header,
                    "error_code": "duplicate_header",
                    "message": f"表头 {header} 重复出现，请删除重复列后重新导入",
                }
            )
        else:
            seen_headers[header] = index
    for required_header in GERMPLASM_REQUIRED_HEADERS:
        if required_header not in headers:
            errors.append(
                {
                    "row_no": header_row_no,
                    "column_name": required_header,
                    "error_code": "missing_required_header",
                    "message": f"缺少必填列 {required_header}",
                }
            )
    return errors, warnings


def _validate_blank_header_columns(
    headers: list[str],
    data_rows: list[tuple[int, tuple]],
    *,
    header_row_no: int,
) -> list[dict]:
    errors: list[dict] = []
    blank_indexes = [index for index, header in enumerate(headers) if not header]
    if not blank_indexes:
        return errors

    for blank_index in blank_indexes:
        for excel_row_no, row_values in data_rows:
            if blank_index >= len(row_values):
                continue
            if _clean_cell(row_values[blank_index]) is None:
                continue
            errors.append(
                {
                    "row_no": header_row_no,
                    "column_name": f"column_{blank_index + 1}",
                    "error_code": "blank_header_with_data",
                    "message": (
                        f"第 {blank_index + 1} 列表头为空，但数据区第 {excel_row_no} 行存在值。"
                        "请为该列提供表头名称，或删除整列空表头数据。"
                    ),
                }
            )
            break
    return errors


def _normalize_dynamic_field_key(source_header: str, *, display_order: int) -> str:
    normalized = re.sub(r"[^0-9A-Za-z]+", "_", (source_header or "").strip()).strip("_").lower()
    if normalized:
        return normalized
    return f"attr_{display_order:03d}"


def _build_field_schema(headers: list[str]) -> tuple[list[dict], list[dict], list[dict]]:
    builtin_fields: list[dict] = []
    dynamic_fields: list[dict] = []
    field_schema: list[dict] = []
    dynamic_order = 0
    for header_index, header in enumerate(headers, start=1):
        if not header:
            continue
        builtin_spec = BUILTIN_FIELD_SPECS.get(header)
        if builtin_spec is not None:
            item = {
                "field_key": builtin_spec["field_key"],
                "field_label": builtin_spec["field_label"],
                "source_header": header,
                "display_order": header_index,
                "data_type": builtin_spec["data_type"],
                "is_builtin": 1,
                "is_dynamic": 0,
            }
            builtin_fields.append(item)
            field_schema.append(item)
            continue

        dynamic_order += 1
        item = {
            "field_key": _normalize_dynamic_field_key(header, display_order=dynamic_order),
            "field_label": header,
            "source_header": header,
            "display_order": dynamic_order,
            "data_type": "string",
            "is_builtin": 0,
            "is_dynamic": 1,
        }
        dynamic_fields.append(item)
        field_schema.append(item)
    return builtin_fields, dynamic_fields, field_schema


def validate_rose_germplasm_file(
    source_path: str,
    *,
    taxonomy_tax_id: int,
    db_existing_accessions: set[str] | None = None,
) -> dict:
    db_existing_accessions = db_existing_accessions or set()
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row_no, headers, data_rows = _load_sheet_headers_and_rows(sheet)
    errors, warnings = _validate_germplasm_headers(headers, header_row_no=header_row_no)
    errors.extend(
        _validate_blank_header_columns(
            headers,
            data_rows,
            header_row_no=header_row_no,
        )
    )
    builtin_fields, dynamic_fields, field_schema = _build_field_schema(headers)
    dynamic_field_map = {item["source_header"]: item["field_key"] for item in dynamic_fields}

    seen_accessions: dict[str, int] = {}
    normalized_rows: list[dict] = []
    preview_rows: list[dict] = []
    valid_rows = 0
    warning_count = len(warnings)

    for excel_row_no, row_values in data_rows:
        row_map = {
            headers[index]: row_values[index] if index < len(row_values) else None
            for index in range(len(headers))
        }
        accession_id = _clean_cell(row_map.get("ID"))
        display_name = _clean_cell(row_map.get("chinese_name"))
        english_name = _clean_cell(row_map.get("english_name"))
        father_accession = _clean_cell(row_map.get("P"))
        mother_accession = _clean_cell(row_map.get("M"))
        father_name = _clean_cell(row_map.get("P_name"))
        mother_name = _clean_cell(row_map.get("M_name"))

        row_errors = []
        row_warnings = []
        if accession_id is None:
            row_errors.append(
                {
                    "row_no": excel_row_no,
                    "column_name": "ID",
                    "error_code": "missing_accession",
                    "message": "种质编号不能为空",
                }
            )
        if display_name is None:
            row_errors.append(
                {
                    "row_no": excel_row_no,
                    "column_name": "chinese_name",
                    "error_code": "missing_display_name",
                    "message": "显示名称不能为空",
                }
            )
        if accession_id:
            if accession_id in seen_accessions:
                row_errors.append(
                    {
                        "row_no": excel_row_no,
                        "column_name": "ID",
                        "error_code": "duplicate_accession",
                        "message": f"种质编号 {accession_id} 在当前文件中重复，首次出现在第 {seen_accessions[accession_id]} 行",
                    }
                )
            else:
                seen_accessions[accession_id] = excel_row_no
            if accession_id in db_existing_accessions:
                row_errors.append(
                    {
                        "row_no": excel_row_no,
                        "column_name": "ID",
                        "error_code": "existing_accession",
                        "message": f"种质编号 {accession_id} 在当前物种下已存在，不能重复导入",
                    }
                )
        if accession_id and father_accession and accession_id == father_accession:
            row_errors.append(
                {
                    "row_no": excel_row_no,
                    "column_name": "P",
                    "error_code": "self_parent_reference",
                    "message": "父本编号不能等于自身编号",
                }
            )
        if accession_id and mother_accession and accession_id == mother_accession:
            row_errors.append(
                {
                    "row_no": excel_row_no,
                    "column_name": "M",
                    "error_code": "self_parent_reference",
                    "message": "母本编号不能等于自身编号",
                }
            )
        if father_accession and mother_accession and father_accession == mother_accession:
            row_warnings.append(
                {
                    "row_no": excel_row_no,
                    "column_name": "P/M",
                    "error_code": "same_parent_accession",
                    "message": "父本和母本编号相同，请确认该记录是否合理",
                }
            )

        attributes = {}
        for header in headers:
            if not header:
                continue
            if header in {"ID", "chinese_name", "english_name", "P", "M", "P_name", "M_name"}:
                continue
            value = _clean_cell(row_map.get(header))
            if value is not None:
                attributes[dynamic_field_map.get(header, header)] = value

        if not row_errors:
            valid_rows += 1
            normalized_row = {
                "taxonomy_tax_id": taxonomy_tax_id,
                "accession_id": accession_id,
                "display_name": display_name,
                "english_name": english_name,
                "father_accession": father_accession,
                "mother_accession": mother_accession,
                "father_name_snapshot": father_name,
                "mother_name_snapshot": mother_name,
                "source_row_no": excel_row_no,
                "attributes_json": attributes,
                "search_text": " ".join(
                    value
                    for value in [
                        accession_id,
                        display_name,
                        english_name,
                        father_accession,
                        mother_accession,
                        " ".join(str(item) for item in attributes.values()) if attributes else None,
                    ]
                    if value
                ),
            }
            normalized_rows.append(normalized_row)
            if len(preview_rows) < 20:
                preview_rows.append(
                    {
                        "accession_id": accession_id,
                        "display_name": display_name,
                        "english_name": english_name,
                        "father_accession": father_accession,
                        "mother_accession": mother_accession,
                        "source_row_no": excel_row_no,
                    }
                )

        errors.extend(row_errors)
        warnings.extend(row_warnings)
        warning_count += len(row_warnings)

    workbook.close()
    return {
        "summary": {
            "passed": len(errors) == 0,
            "template_profile": CANONICAL_TEMPLATE_PROFILE,
            "taxonomy_tax_id": taxonomy_tax_id,
            "total_rows": len(normalized_rows) + len({item["row_no"] for item in errors if item.get("row_no", 0) > 1}),
            "valid_rows": valid_rows,
            "error_rows": len({item["row_no"] for item in errors if item.get("row_no", 0) > 1}),
            "warning_rows": warning_count,
        },
        "errors": errors,
        "warnings": warnings,
        "preview_rows": preview_rows,
        "builtin_fields": builtin_fields,
        "dynamic_fields": dynamic_fields,
        "field_schema_preview": field_schema,
        "normalized_rows": normalized_rows,
    }


def validate_germplasm_file(
    template_profile: str,
    source_path: str,
    *,
    taxonomy_tax_id: int,
    db_existing_accessions: set[str] | None = None,
) -> dict:
    normalized_profile = normalize_template_profile(template_profile)
    if normalized_profile == CANONICAL_TEMPLATE_PROFILE:
        return validate_rose_germplasm_file(
            source_path,
            taxonomy_tax_id=taxonomy_tax_id,
            db_existing_accessions=db_existing_accessions,
        )
    raise ValueError(f"unsupported template profile: {template_profile}")


def remove_validation_bundle(token: str) -> None:
    bundle_dir = ensure_validation_root() / token
    if bundle_dir.exists():
        shutil.rmtree(bundle_dir, ignore_errors=True)
